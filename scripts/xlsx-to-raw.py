# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""Dump every worksheet of "60 acres 2025.xlsx" into scripts/raw-tabs/NN.txt in the
exact `TAB: <name>` / `Row N: [...]` format that parse-tabs.mjs already consumes.
The local Excel file has the FULL grid (no 50-row API cap), so this replaces the
unreliable read_sheet_values dumping entirely."""
import os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "60 acres 2025.xlsx")
OUT = os.path.join(HERE, "raw-tabs")
os.makedirs(OUT, exist_ok=True)

# Wipe any stale dumps so we don't stitch old truncated files with new ones.
for f in os.listdir(OUT):
    if f.endswith(".txt"):
        os.remove(os.path.join(OUT, f))

wb = load_workbook(XLSX, read_only=True, data_only=True)
print(f"sheets ({len(wb.sheetnames)}):")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i:02d}  {name!r}")

for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    lines = [f"TAB: {name}", ""]
    for r, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Normalize each cell to a trimmed string; None -> ''.
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            elif isinstance(v, float) and v.is_integer():
                # Excel stores the slot numbers 1,2,3.. as floats; render "1" not
                # "1.0" so parse-tabs.mjs's numRe (/^\d{1,2}$/) matches them.
                cells.append(str(int(v)))
            else:
                cells.append(str(v).strip())
        # Drop fully-trailing-empty cells so lines aren't padded to the sheet width.
        while cells and cells[-1] == "":
            cells.pop()
        if not cells:
            continue
        # Emit in the Row N: ['a', 'b', ...] shape parseRow expects. Escape ' and \.
        rendered = ", ".join("'" + c.replace("\\", "\\\\").replace("'", "\\'") + "'" for c in cells)
        lines.append(f"Row {r}: [{rendered}]")
    with open(os.path.join(OUT, f"{i:02d}.txt"), "w") as fh:
        fh.write("\n".join(lines) + "\n")
    print(f"wrote {i:02d}.txt  ({name!r}, {len(lines)-2} non-empty rows)")

wb.close()
